'''
createPDF()
txt to pdf(txt)
docx to pdf(docx)
docx to txt(docx)
pdf to docx(pdf)
ppt to pdf(ppt)
pdf to txt(txt)
xlsx to pdf(xlsx)
unZip(zipFile)
CompressToZip(folder)
imageView(img)
compressImage(img)


next update:
csv to xlsx(csv)
imageResize(width, height)
sendMail()
uploadToDrive(file)
'''

def createPDF(pdfFileName, textToBeEntered):
    # pdfFileName is the name in which the user would like to have the PDF file created. 
    from fpdf import FPDF
    # import the FPDF function 
    pdf = FPDF()
    # for convinience purposes 
    pdf.add_page()
    # add a page to the PDF 
    pdf.set_font('Arial', size = 15)
    # setting font and font size 
    pdf.cell(200, 10, txt = textToBeEntered, align = 'C')
    # Add the text to the PDF by adding the cell. 
    pdf.output(f"{pdfFileName}.pdf")
    # output the PDF with the fileName provided by the user with the PDF extension

def txtToPDF(txtFileName):
    from fpdf import FPDF 
    # import the FPDF function 
    pdf = FPDF()
    # for convinience purposes 
    pdf.add_page() 
    # add a page to the PDF 
    pdf.set_font('Arial', size = 15)
    # setting font and font size 
    file = open(f"{txtFileName}.txt", r)
    # open the file in read mode 
    for x in file: 
        pdf.cell(200, 10, txt = x, ln = 1, align = 'C') 
    # read all the text in the file and add it in the pdf  
    pdf.output(f"{txtFileName}.pdf")
    # output the PDF with the same name and .pdf extension 

def docxToPDF(docxFileName):
    from docx2pdf import convert
    if '/' in docxFileName:
        convert(f"{docxFileName}")
        # if the given name is a folder name, then it would contain a /. Hence, we use the convert method used for folder. 
        # this function will convert all the docx files inside the folder into PDF and store them again inside the same folder. 
    else:
        convert(f"{docxFileName}.docx", f"{docxFileName}.pdf")
        # if the given name is a file name, then it will not contain a / and hence, we use the convert method used for files. 

def docxToTxt(docxFileName):
    import docx2txt
    # importing docx2txt module 
    textInsideDocx = docx2txt.process(f"{docxFileName}.docx")
    # process and save the text inside the docx file inside a variable. 

    with open(f"{docxFileName}.txt", 'w') as txtFile:
        # create a txt file with the same name of the docx file 
        print(textInsideDocx, file=txtFile)
        # add the text inside the created txt file 

def pdfToDocx(pdfFileName):
    from pdf2docx import parse
    # import the parse function from pdf2docx 
    docxFileName = pdfFileName + ".docx"
    # save the docx file name in a variable
    pdfFileName = pdfFileName + ".pdf"
    # add the .pdf extension to the pdfFileName. 
    parse(pdfFileName, docxFileName, start = 0, end = 1)
    # use the parse function to convert pdf to docx 

def pptToPDF(pptFileName, pdfFileName, formatType = 32):
    import comtypes.client
    # comtypes is available only for windows 
    powerpoint = comtypes.client.CreateObject("Powerpoint.Application")
    powerpoint.Visible = 1

    if pdfFileName[-3:] != 'pdf':
        pdfFileName = pdfFileName + ".pdf"
    deck = powerpoint.Presentations.Open(pptFileName)
    deck.SaveAs(pdfFileName, formatType) 
    # formatType = 32 for ppt to pdf
    deck.Close()
    powerpoint.Quit()

def pdfToTxt(pdfFileName):
    import PyPDF2
    # import the PyPDF2 module 
    pdfFileObject = open(f'{pdfFileName}.pdf', 'rb')
    # opening a PDF File Object 
    pdfReader = PyPDF2.PdfFileReader(pdfFileObject)
    # initiating a PDF Reader 
    numberOFPagesInPDF = pdfReader.numPages
    # gettting the Number of Pages in the PDF 
    with open(f"{pdfFileName}.txt", 'a') as txtFile:
        # creating a new txt file 
        for i in range(0, numberOFPagesInPDF-1):
            currentPageObj = pdfReader.getPage(i)
            # creating a PDF Page object 
            textInPage = currentPageObj.extractText()
            # extracting the text from the current PDF Page Object 
            txtFile.write(textInPage)
            # appending the text in text file 
    pdfFileObject.close()
    # closing the PDF File object 
        
def createExcelFile(excelFileName): # Incomplete 
    import openpyxl
    currentWorkbook = openpyxl.Workbook()
    currentWorkSheet = currentWorkbook.active()

def excelToPDF(excelFileName):
    import pandas 
    import pdfkit

    excelFileObject = pandas.read_excel(f"{excelFileName}.xlsx")
    # creating an excel file object 
    excelFileObject.to_html(f"{excelFileName}.html")
    # converting the excel file into an html file 
    pdfkit.from_file(f"{excelFileName}.html", f"{excelFileName}.pdf")
    # converting html file to pdf 

def unZipFile(zipFileName):
    from zipfile import ZipFile
    with ZipFile(file_name, 'r') as zip: 
        # opening the zip file in read mode 
        zip.printdir() 
        # printing all the contents of the zip file 
    
        print('Extracting all the files now...') 
        zip.extractall() 
        # extracting all the files 
        print('Done!') 

def compressToZip(folderName):
    from zipfile import ZipFile
    with ZipFile(f'{folderName}.zip', mode='w'):
        ZipFile.write(folderName)

def openImage(imageName):
    from PIL import Image
    currentImage = Image.open(imageName)
    currentImage.show()

def compressImage(imageName):
    from PIL import Image
    currentImage = Image.open(imageName)
    # opening the Image file 
    currentImage.save("Compressed_"+imageName, quality = 5)
    # compressing the image file 

# WORK ON IMAGE FORMATS 
